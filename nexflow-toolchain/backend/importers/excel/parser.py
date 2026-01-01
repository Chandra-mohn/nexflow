"""
Excel workbook parser for service domain data models.

Parses Excel workbooks with the following sheets:
- ServiceDomain: Domain metadata
- Entities: Schema/type definitions
- Attributes: Field definitions
- Relationships: Parent-child composition
- Enumerations: Enum values
"""

from pathlib import Path
from typing import Optional
import logging

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    raise ImportError(
        "openpyxl is required for Excel import. Install with: pip install openpyxl"
    )

from .models import (
    ServiceDomain,
    Entity,
    Attribute,
    Relationship,
    EnumValue,
    EntityQualifier,
    PersistencePattern,
    EntityStereotype,
    ObjectType,
    RelationshipType,
)

logger = logging.getLogger(__name__)


class ExcelParseError(Exception):
    """Error during Excel parsing."""


class ExcelParser:
    """Parse Excel workbook into ServiceDomain model."""

    REQUIRED_SHEETS = ["ServiceDomain", "Entities", "Attributes", "Relationships", "Enumerations"]

    def __init__(self, file_path: Path):
        self.file_path = file_path
        self.workbook = None
        self.warnings: list[str] = []
        self.errors: list[str] = []

    def parse(self) -> ServiceDomain:
        """Parse the Excel workbook and return a ServiceDomain."""
        logger.info(f"Parsing Excel workbook: {self.file_path}")

        self.workbook = load_workbook(self.file_path, read_only=True, data_only=True)

        # Validate required sheets
        self._validate_sheets()

        # Parse each section
        domain_name = self._parse_service_domain()
        entities = self._parse_entities()
        attributes = self._parse_attributes()
        relationships = self._parse_relationships()
        enum_values = self._parse_enumerations()

        # Build the domain model
        domain = ServiceDomain(
            name=domain_name,
            source_file=str(self.file_path),
            entities={e.entity_name: e for e in entities},
            relationships=relationships,
            enum_values=enum_values,
        )

        # Attach attributes to entities
        self._attach_attributes(domain, attributes)

        # Attach relationships to entities
        self._attach_relationships(domain)

        # Attach enum values to attributes
        self._attach_enums(domain)

        # Classify entities
        self._classify_entities(domain)

        self.workbook.close()

        if self.errors:
            raise ExcelParseError(f"Parse errors: {'; '.join(self.errors)}")

        return domain

    def _validate_sheets(self):
        """Validate that all required sheets exist."""
        sheet_names = self.workbook.sheetnames
        for required in self.REQUIRED_SHEETS:
            if required not in sheet_names:
                self.errors.append(f"MISSING_SHEET: Required sheet '{required}' not found")

    def _get_sheet(self, name: str) -> Optional[Worksheet]:
        """Get a worksheet by name."""
        if name in self.workbook.sheetnames:
            return self.workbook[name]
        return None

    def _get_column_index(self, headers: list[str], column_name: str) -> Optional[int]:
        """Find column index by header name (case-insensitive)."""
        column_lower = column_name.lower()
        for i, header in enumerate(headers):
            if header and header.lower() == column_lower:
                return i
        return None

    def _get_cell_value(self, row: tuple, index: Optional[int], default: str = "") -> str:
        """Get cell value at index, returning default if None or out of bounds."""
        if index is None or index >= len(row):
            return default
        value = row[index]
        if value is None:
            return default
        return str(value).strip()

    def _parse_service_domain(self) -> str:
        """Parse ServiceDomain sheet to get domain name."""
        sheet = self._get_sheet("ServiceDomain")
        if not sheet:
            return "Unknown"

        # Look for row where Name = "Service Domain Name"
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 2:
                name_col = str(row[0]).strip() if row[0] else ""
                value_col = str(row[1]).strip() if row[1] else ""
                if name_col.lower() == "service domain name":
                    return value_col

        self.warnings.append("Service Domain Name not found, using filename")
        return self.file_path.stem

    def _parse_entities(self) -> list[Entity]:
        """Parse Entities sheet."""
        sheet = self._get_sheet("Entities")
        if not sheet:
            return []

        entities = []
        rows = list(sheet.iter_rows(values_only=True))

        if len(rows) < 2:
            return []

        # Get headers from first row
        headers = [str(h).strip() if h else "" for h in rows[0]]

        # Column indices
        idx_entity_name = self._get_column_index(headers, "EntityName")
        idx_entity_code = self._get_column_index(headers, "EntityCode")
        idx_object_type = self._get_column_index(headers, "ObjectType")
        idx_collection_name = self._get_column_index(headers, "CollectionName")
        idx_pattern = self._get_column_index(headers, "Pattern")
        idx_stereotype = self._get_column_index(headers, "Stereotype")
        idx_qualifier = self._get_column_index(headers, "EntityQualifier")
        idx_bim_entity = self._get_column_index(headers, "BIMEntityName")
        idx_comment = self._get_column_index(headers, "Comment")

        for row in rows[1:]:
            if not row or not row[0]:
                continue

            entity_name = self._get_cell_value(row, idx_entity_name)
            if not entity_name:
                continue

            collection_name = self._get_cell_value(row, idx_collection_name)
            if not collection_name:
                collection_name = entity_name
                self.warnings.append(
                    f"MISSING_COLLECTION_NAME: Entity '{entity_name}' has no CollectionName, using EntityName"
                )

            entity = Entity(
                entity_name=entity_name,
                entity_code=self._get_cell_value(row, idx_entity_code),
                collection_name=collection_name,
                object_type=self._parse_object_type(self._get_cell_value(row, idx_object_type)),
                pattern=self._parse_pattern(self._get_cell_value(row, idx_pattern)),
                stereotype=self._parse_stereotype(self._get_cell_value(row, idx_stereotype)),
                qualifier=self._parse_qualifier(self._get_cell_value(row, idx_qualifier)),
                bim_entity_name=self._get_cell_value(row, idx_bim_entity) or None,
                comment=self._get_cell_value(row, idx_comment) or None,
            )
            entities.append(entity)

        logger.info(f"Parsed {len(entities)} entities")
        return entities

    def _parse_attributes(self) -> list[Attribute]:
        """Parse Attributes sheet."""
        sheet = self._get_sheet("Attributes")
        if not sheet:
            return []

        attributes = []
        rows = list(sheet.iter_rows(values_only=True))

        if len(rows) < 2:
            return []

        headers = [str(h).strip() if h else "" for h in rows[0]]

        # Column indices
        idx_entity_name = self._get_column_index(headers, "EntityName")
        idx_entity_code = self._get_column_index(headers, "EntityCode")
        idx_name = self._get_column_index(headers, "Name")
        idx_code = self._get_column_index(headers, "Code")
        idx_datatype = self._get_column_index(headers, "Datatype")
        idx_length = self._get_column_index(headers, "Length")
        idx_mandatory = self._get_column_index(headers, "MandatoryFlag")
        idx_primary = self._get_column_index(headers, "PrimaryIdentifierFlag")
        idx_foreign = self._get_column_index(headers, "ForeignIdentifierFlag")
        idx_foreign_parent = self._get_column_index(headers, "ForeignIdentifierParentEntityName")
        idx_foreign_parent_code = self._get_column_index(headers, "ForeignIdentifierParentEntityCode")
        idx_json_name = self._get_column_index(headers, "JSONName")
        idx_pii = self._get_column_index(headers, "PIIPCIIndicator")
        idx_stereotype = self._get_column_index(headers, "Stereotype")
        idx_identifier_type = self._get_column_index(headers, "IdentifierType")
        idx_domain = self._get_column_index(headers, "Domain")
        idx_comment = self._get_column_index(headers, "Comment")
        idx_copybook = self._get_column_index(headers, "PrimaryCopybookFieldName")

        for row in rows[1:]:
            if not row or not row[0]:
                continue

            entity_name = self._get_cell_value(row, idx_entity_name)
            attr_name = self._get_cell_value(row, idx_name)
            if not entity_name or not attr_name:
                continue

            attr = Attribute(
                entity_name=entity_name,
                entity_code=self._get_cell_value(row, idx_entity_code),
                name=attr_name,
                code=self._get_cell_value(row, idx_code),
                datatype=self._get_cell_value(row, idx_datatype),
                length=self._get_cell_value(row, idx_length) or None,
                mandatory=self._parse_boolean(self._get_cell_value(row, idx_mandatory)),
                primary_identifier=self._parse_boolean(self._get_cell_value(row, idx_primary)),
                foreign_identifier=self._parse_boolean(self._get_cell_value(row, idx_foreign)),
                foreign_parent_entity=self._get_cell_value(row, idx_foreign_parent) or None,
                foreign_parent_code=self._get_cell_value(row, idx_foreign_parent_code) or None,
                json_name=self._get_cell_value(row, idx_json_name) or None,
                pii_pci_indicator=self._get_cell_value(row, idx_pii) or None,
                stereotype=self._parse_stereotype(self._get_cell_value(row, idx_stereotype)),
                identifier_type=self._get_cell_value(row, idx_identifier_type) or None,
                domain=self._get_cell_value(row, idx_domain) or None,
                comment=self._get_cell_value(row, idx_comment) or None,
                copybook_field=self._get_cell_value(row, idx_copybook) or None,
            )
            attributes.append(attr)

        logger.info(f"Parsed {len(attributes)} attributes")
        return attributes

    def _parse_relationships(self) -> list[Relationship]:
        """Parse Relationships sheet."""
        sheet = self._get_sheet("Relationships")
        if not sheet:
            return []

        relationships = []
        rows = list(sheet.iter_rows(values_only=True))

        if len(rows) < 2:
            return []

        headers = [str(h).strip() if h else "" for h in rows[0]]

        idx_name = self._get_column_index(headers, "RelationshipName")
        idx_code = self._get_column_index(headers, "RelationshipCode")
        idx_parent = self._get_column_index(headers, "ParentEntity")
        idx_child = self._get_column_index(headers, "ChildEntity")
        idx_type = self._get_column_index(headers, "RelationshipType")
        idx_join = self._get_column_index(headers, "JoinAttributes")
        idx_stereotype = self._get_column_index(headers, "Stereotype")
        idx_comment = self._get_column_index(headers, "Comment")

        for row in rows[1:]:
            if not row:
                continue

            parent = self._get_cell_value(row, idx_parent)
            child = self._get_cell_value(row, idx_child)
            if not parent or not child:
                continue

            rel = Relationship(
                name=self._get_cell_value(row, idx_name),
                code=self._get_cell_value(row, idx_code),
                parent_entity=parent,
                child_entity=child,
                relationship_type=self._parse_relationship_type(self._get_cell_value(row, idx_type)),
                join_attributes=self._get_cell_value(row, idx_join) or None,
                stereotype=self._get_cell_value(row, idx_stereotype),
                comment=self._get_cell_value(row, idx_comment) or None,
            )
            relationships.append(rel)

        logger.info(f"Parsed {len(relationships)} relationships")
        return relationships

    def _parse_enumerations(self) -> list[EnumValue]:
        """Parse Enumerations sheet."""
        sheet = self._get_sheet("Enumerations")
        if not sheet:
            return []

        enum_values = []
        rows = list(sheet.iter_rows(values_only=True))

        if len(rows) < 2:
            return []

        headers = [str(h).strip() if h else "" for h in rows[0]]

        idx_entity = self._get_column_index(headers, "EntityName")
        idx_attribute = self._get_column_index(headers, "AttributeName")
        idx_value = self._get_column_index(headers, "Value")
        idx_label = self._get_column_index(headers, "Label")

        for row in rows[1:]:
            if not row:
                continue

            entity = self._get_cell_value(row, idx_entity)
            attribute = self._get_cell_value(row, idx_attribute)
            value = self._get_cell_value(row, idx_value)
            if not entity or not attribute or not value:
                continue

            enum_val = EnumValue(
                entity_name=entity,
                attribute_name=attribute,
                value=value,
                label=self._get_cell_value(row, idx_label) or value,
            )
            enum_values.append(enum_val)

        logger.info(f"Parsed {len(enum_values)} enum values")
        return enum_values

    def _attach_attributes(self, domain: ServiceDomain, attributes: list[Attribute]):
        """Attach attributes to their parent entities."""
        for attr in attributes:
            entity = domain.get_entity_by_name(attr.entity_name)
            if entity:
                entity.attributes.append(attr)
            else:
                self.warnings.append(
                    f"ORPHAN_ATTRIBUTE: Attribute '{attr.name}' references unknown entity '{attr.entity_name}'"
                )

    def _attach_relationships(self, domain: ServiceDomain):
        """Attach child relationships to parent entities."""
        for rel in domain.relationships:
            parent = domain.get_entity_by_name(rel.parent_entity)
            if parent:
                parent.child_relationships.append(rel)
            else:
                self.warnings.append(
                    f"ORPHAN_RELATIONSHIP: Relationship '{rel.name}' references unknown parent '{rel.parent_entity}'"
                )

    def _attach_enums(self, domain: ServiceDomain):
        """Attach enum values to entity attributes."""
        for enum_val in domain.enum_values:
            entity = domain.get_entity_by_name(enum_val.entity_name)
            if entity:
                if enum_val.attribute_name not in entity.enum_values:
                    entity.enum_values[enum_val.attribute_name] = []
                entity.enum_values[enum_val.attribute_name].append(enum_val)

    def _classify_entities(self, domain: ServiceDomain):
        """Classify entities as managed or external."""
        for entity in domain.entities.values():
            if entity.is_managed():
                domain.managed_entities.append(entity.entity_name)
            if entity.is_shortcut():
                domain.external_references.append(entity.entity_name)

    def _parse_object_type(self, value: str) -> ObjectType:
        """Parse ObjectType value."""
        if value.lower() == "shortcut entity":
            return ObjectType.SHORTCUT
        return ObjectType.ENTITY

    def _parse_pattern(self, value: str) -> PersistencePattern:
        """Parse Pattern value."""
        value_lower = value.lower()
        if value_lower == "master":
            return PersistencePattern.MASTER
        elif value_lower == "event":
            return PersistencePattern.EVENT
        elif value_lower == "ledger":
            return PersistencePattern.LEDGER
        return PersistencePattern.NONE

    def _parse_stereotype(self, value: str) -> EntityStereotype:
        """Parse Stereotype value."""
        value_lower = value.lower()
        if value_lower == "array":
            return EntityStereotype.ARRAY
        elif value_lower == "enum":
            return EntityStereotype.ENUM
        return EntityStereotype.NONE

    def _parse_qualifier(self, value: str) -> EntityQualifier:
        """Parse EntityQualifier value."""
        value_upper = value.upper()
        if value_upper == "BOM":
            return EntityQualifier.BOM
        elif value_upper == "BQ":
            return EntityQualifier.BQ
        elif value_upper == "CR":
            return EntityQualifier.CR
        return EntityQualifier.NONE

    def _parse_relationship_type(self, value: str) -> RelationshipType:
        """Parse RelationshipType value."""
        if value == "1:n" or value == "1:N":
            return RelationshipType.ONE_TO_MANY
        return RelationshipType.ONE_TO_ONE

    def _parse_boolean(self, value: str) -> bool:
        """Parse boolean flag value."""
        return value.upper() in ("Y", "YES", "TRUE", "1")
